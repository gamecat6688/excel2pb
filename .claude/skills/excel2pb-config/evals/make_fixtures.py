"""生成评测用的 xlsx 夹具。"""
import os
import openpyxl
from openpyxl.comments import Comment

HERE = os.path.dirname(os.path.abspath(__file__))
FILES = os.path.join(HERE, "files")
os.makedirs(FILES, exist_ok=True)


def header(ws, cols):
    """cols: list of (name, type, filter, desc, comment_or_None)"""
    for c, (n, t, f, d, cm) in enumerate(cols, start=1):
        ws.cell(row=1, column=c, value=n)
        ws.cell(row=2, column=c, value=t)
        ws.cell(row=3, column=c, value=f)
        ws.cell(row=4, column=c, value=d)
        if cm:
            ws.cell(row=1, column=c).comment = Comment(cm, "designer")


def put_rows(ws, rows, start=5):
    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            ws.cell(row=start + i, column=j, value=v)


# ---- 夹具1: 物品.xlsx —— 设计新表(eval-0)时作为外键目标 ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Item"
header(ws, [
    ("ID",   "pk int32", "cs", "道具ID", None),
    ("Name", "i18n",     "cs", "道具名", None),
])
put_rows(ws, [[1001, "长剑"], [1002, "圆盾"], [1003, "生命药水"]])
wb.save(os.path.join(FILES, "物品.xlsx"))

# ---- 夹具2: 商店.xlsx —— 填表(eval-1)。结构齐全但商店无数据 ----
wb = openpyxl.Workbook()
# Item
ws = wb.active
ws.title = "Item"
header(ws, [
    ("ID",   "pk int32", "cs", "道具ID", None),
    ("Name", "i18n",     "cs", "道具名", None),
])
put_rows(ws, [[2001, "初级药水"], [2002, "魔法宝石"]])
# Reward (嵌套结构定义)
ws = wb.create_sheet("Reward")
header(ws, [
    ("ItemID", "pk int32", "cs", "道具ID", None),
    ("Count",  "int32",    "cs", "数量",   None),
])
# Shop (待填)
ws = wb.create_sheet("Shop")
header(ws, [
    ("ID",      "pk int32",        "cs", "商店项ID", None),
    ("Name",    "i18n",            "cs", "商品名",   None),
    ("Price",   "int32",           "cs", "价格",     None),
    ("Rewards", "repeated Reward", "s",  "购买获得", "fk:Reward.ItemID=Item.ID"),
])
wb.save(os.path.join(FILES, "商店.xlsx"))

# ---- 夹具3: 怪物.xlsx —— 校验(eval-2)。故意埋错 ----
wb = openpyxl.Workbook()
# Item (fk 目标, 合法)
ws = wb.active
ws.title = "Item"
header(ws, [
    ("ID",   "pk int32", "cs", "道具ID", None),
    ("Name", "i18n",     "cs", "道具名", None),
])
put_rows(ws, [[1001, "金币"], [1002, "经验书"]])
# Monster: 埋 3 个错——主键重复、外键失效、时间戳格式错
ws = wb.create_sheet("Monster")
header(ws, [
    ("ID",         "pk int32",   "cs", "怪物ID",   None),
    ("Name",       "string",     "cs", "怪物名",   None),
    ("DropItemID", "int32",      "cs", "掉落道具", "fk:Item.ID"),
    ("SpawnTime",  "timestamp",  "s",  "刷新时间", None),
])
put_rows(ws, [
    [5, "史莱姆", 1001, "2025-03-01 08:00:00"],
    [5, "哥布林", 9999, "2025/03/01 09:00"],   # ID 重复(5), DropItemID=9999 不存在, 时间格式错
    [7, "巨龙",   1002, "2025-03-02 10:00:00"],
])
# Tips: i18n 字段但没有主键 —— i18n 缺主键错
ws = wb.create_sheet("Tips")
header(ws, [
    ("Content", "i18n", "cs", "提示文本", None),
])
put_rows(ws, [["小心脚下"], ["记得补给"]])
wb.save(os.path.join(FILES, "怪物.xlsx"))

print("fixtures written to", FILES)
for f in sorted(os.listdir(FILES)):
    print(" -", f)
