#!/usr/bin/env python3
"""Apply readable, content-aware formatting to excel2pb workbooks."""

from __future__ import annotations

import argparse
import math
import unicodedata
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


DEFAULT_MIN_WIDTH = 10.0
DEFAULT_MAX_WIDTH = 48.0
DEFAULT_PADDING = 2.0
DEFAULT_LINE_HEIGHT = 18.0
MAX_ROW_HEIGHT = 180.0


def display_width(value: object) -> int:
    """Approximate Excel display width, counting wide CJK characters as two."""
    text = "" if value is None else str(value)
    return sum(2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1 for char in text)


def cell_line_widths(value: object) -> list[int]:
    text = "" if value is None else str(value)
    return [display_width(line) for line in text.splitlines() or [""]]


def discover_workbooks(paths: list[Path]) -> list[Path]:
    workbooks: set[Path] = set()
    for path in paths:
        if path.is_dir():
            workbooks.update(
                candidate.resolve()
                for candidate in path.rglob("*.xlsx")
                if not candidate.name.startswith("~")
            )
        elif path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~"):
            workbooks.add(path.resolve())
        else:
            raise ValueError(f"Not an .xlsx workbook or directory: {path}")
    if not workbooks:
        raise ValueError("No .xlsx workbooks found")
    return sorted(workbooks)


def wrapped_line_count(value: object, available_width: int) -> int:
    return sum(
        max(1, math.ceil(line_width / max(1, available_width)))
        for line_width in cell_line_widths(value)
    )


def format_sheet(sheet, min_width: float, max_width: float, padding: float) -> int:
    column_widths: dict[int, float] = {}
    for column in range(1, sheet.max_column + 1):
        widest = max(
            max(cell_line_widths(sheet.cell(row=row, column=column).value))
            for row in range(1, sheet.max_row + 1)
        )
        width = min(max_width, max(min_width, widest + padding))
        sheet.column_dimensions[get_column_letter(column)].width = width
        column_widths[column] = width

    formatted_cells = 0
    for row in range(1, sheet.max_row + 1):
        row_lines = 1
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            alignment = copy(cell.alignment)
            if row <= 4:
                alignment.horizontal = "center"
                alignment.vertical = "center"
            available_width = max(1, int(column_widths[column] - padding))
            needs_wrap = row == 4 or len(cell_line_widths(cell.value)) > 1 or any(
                line_width > available_width for line_width in cell_line_widths(cell.value)
            )
            if needs_wrap:
                alignment.wrap_text = True
                alignment.vertical = "center"
            cell.alignment = alignment
            if alignment.wrap_text:
                row_lines = max(row_lines, wrapped_line_count(cell.value, available_width))
            formatted_cells += 1

        if row_lines > 1:
            current_height = sheet.row_dimensions[row].height or DEFAULT_LINE_HEIGHT
            sheet.row_dimensions[row].height = min(
                MAX_ROW_HEIGHT,
                max(current_height, DEFAULT_LINE_HEIGHT * row_lines),
            )

    return formatted_cells


def format_workbook(path: Path, min_width: float, max_width: float, padding: float) -> tuple[int, int]:
    workbook = load_workbook(path)
    sheet_count = 0
    cell_count = 0
    for sheet in workbook.worksheets:
        sheet_count += 1
        cell_count += format_sheet(sheet, min_width, max_width, padding)
    workbook.save(path)
    return sheet_count, cell_count


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("paths", nargs="+", type=Path, help="Workbook files or directories to format")
    parser.add_argument("--min-width", type=float, default=DEFAULT_MIN_WIDTH)
    parser.add_argument("--max-width", type=float, default=DEFAULT_MAX_WIDTH)
    parser.add_argument("--padding", type=float, default=DEFAULT_PADDING)
    args = parser.parse_args()

    if args.min_width <= 0 or args.max_width < args.min_width or args.padding < 0:
        parser.error("widths must be positive, max-width must be >= min-width, and padding must be non-negative")

    workbook_count = 0
    sheet_count = 0
    cell_count = 0
    for path in discover_workbooks(args.paths):
        sheets, cells = format_workbook(path, args.min_width, args.max_width, args.padding)
        workbook_count += 1
        sheet_count += sheets
        cell_count += cells

    print(
        f"Formatted {workbook_count} workbook(s), {sheet_count} sheet(s), "
        f"and {cell_count} populated-range cell(s)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
